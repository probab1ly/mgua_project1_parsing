import json
import threading
import logging
import webbrowser
import hashlib
import hmac
import os
import secrets
import ssl
import smtplib
import time
import urllib.parse
from collections import defaultdict, deque
from datetime import datetime, timezone
from email.message import EmailMessage
from functools import wraps
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

from scraper import (
    fetch_all_articles,
    get_stats,
    ALL_SOURCES,
    is_ai_relevant,
    is_allowed_source_url,
    has_explicit_ai_signal,
    has_primary_ai_signal,
    is_official_ai_law_page,
    is_monitoring_relevant,
    parse_date_text,
    is_recent_date,
    legal_priority,
    legal_update_type,
    to_datetime,
    deduplicate_articles,
    attach_analysis,
    passes_article_filter,
    relevance_score,
    OFFICIAL_PAGE_SOURCES,
    clean_html,
    date_temporal_status,
)


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

SECRET_KEY_FILE = Path(".ai_monitor_secret")


def load_secret_key() -> str:
    configured = os.environ.get("AI_MONITOR_SECRET_KEY", "").strip()
    if configured:
        if len(configured) < 32:
            raise RuntimeError("AI_MONITOR_SECRET_KEY must contain at least 32 characters")
        return configured
    if SECRET_KEY_FILE.exists():
        stored = SECRET_KEY_FILE.read_text(encoding="utf-8").strip()
        if len(stored) >= 32:
            return stored
    generated = secrets.token_urlsafe(48)
    SECRET_KEY_FILE.write_text(generated, encoding="utf-8")
    try:
        os.chmod(SECRET_KEY_FILE, 0o600)
    except OSError:
        pass
    logger.warning("AI_MONITOR_SECRET_KEY is not set; generated private local key in %s", SECRET_KEY_FILE)
    return generated


app = Flask(__name__)
app.secret_key = load_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    SESSION_COOKIE_SECURE=os.environ.get("AI_MONITOR_HTTPS", "").lower() in {"1", "true", "yes"},
    MAX_CONTENT_LENGTH=64 * 1024,
    MAX_FORM_MEMORY_SIZE=64 * 1024,
    MAX_FORM_PARTS=20,
)

CACHE_FILE = Path("cache.json")
CACHE_TEMP_FILE = Path("cache.json.tmp")
USERS_FILE = Path("users.json")
EMAIL_CONFIG_FILE = Path("email_config.json")
HISTORY_DIR = Path("update_history")
HISTORY_INDEX_FILE = HISTORY_DIR / "history.json"
REFRESH_LOCK_FILE = Path(".refresh.lock")
REFRESH_LOCK_STALE_SECONDS = 15 * 60
CACHE_VERSION = 13
CACHE_LOCK = threading.Lock()
HISTORY_LOCK = threading.RLock()
LOGIN_RATE_LOCK = threading.Lock()
LOGIN_ATTEMPTS: dict[str, deque[float]] = defaultdict(deque)
LOGIN_RATE_WINDOW_SECONDS = 15 * 60
LOGIN_RATE_MAX_ATTEMPTS = 10
WEEKLY_REFRESH_INTERVAL_SECONDS = 7 * 24 * 60 * 60

USERS: dict[str, str] = {}
_cache_file_mtime = 0.0

_cache: dict = {
    "articles": [],
    "stats": {},
    "last_updated": None,
    "sources_count": len(ALL_SOURCES),
    "loading": False,
    "diagnostics": [],
    "last_auto_refresh": None,
    "refresh_owner": None,
    "refresh_status": "idle",
    "refresh_error": None,
    "refresh_finished_at": None,
}



@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=(), payment=(), usb=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; "
        "font-src 'self'; script-src 'self' 'unsafe-inline'; connect-src 'self'; "
        "frame-ancestors 'none'; base-uri 'self'; form-action 'self'",
    )
    if request.is_secure:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    response.headers.pop("Access-Control-Allow-Origin", None)
    return response


def csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def csrf_is_valid() -> bool:
    supplied = request.headers.get("X-CSRF-Token") or request.form.get("_csrf_token", "")
    expected = session.get("_csrf_token", "")
    return bool(supplied and expected and hmac.compare_digest(supplied, expected))


def client_ip() -> str:
    return request.remote_addr or "unknown"


def login_rate_limited() -> bool:
    now = time.monotonic()
    with LOGIN_RATE_LOCK:
        attempts = LOGIN_ATTEMPTS[client_ip()]
        while attempts and now - attempts[0] > LOGIN_RATE_WINDOW_SECONDS:
            attempts.popleft()
        return len(attempts) >= LOGIN_RATE_MAX_ATTEMPTS


def record_login_failure() -> None:
    with LOGIN_RATE_LOCK:
        LOGIN_ATTEMPTS[client_ip()].append(time.monotonic())


def clear_login_failures() -> None:
    with LOGIN_RATE_LOCK:
        LOGIN_ATTEMPTS.pop(client_ip(), None)


def safe_next_url(value: str | None) -> str:
    if not value:
        return url_for("index")
    parsed = urllib.parse.urlsplit(value)
    if parsed.scheme or parsed.netloc or not value.startswith("/") or value.startswith("//"):
        return url_for("index")
    return value


def neutralize_spreadsheet_value(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value



OFFICIAL_SOURCE_NAMES = {source["name"] for source in OFFICIAL_PAGE_SOURCES}
SOURCE_BY_NAME = {source["name"]: source for source in ALL_SOURCES}


def validate_and_analyze_article(article: dict) -> dict | None:
    source_name = article.get("source", "")
    source = SOURCE_BY_NAME.get(source_name, {})
    title = source.get("title") or clean_html(article.get("title", ""))
    summary = clean_html(article.get("summary", ""))
    link = article.get("link", "")
    if not title or not is_allowed_source_url(link):
        return None
    if date_temporal_status(article.get("date", "")) == "future":
        return None
    enriched_text = f"{summary} {source.get('description', '')}".strip()
    if source_name == "Federal Register (официальный API)":
        relevant = has_primary_ai_signal(title, "") and passes_article_filter(title, summary)
    elif source_name in OFFICIAL_SOURCE_NAMES:
        relevant = has_primary_ai_signal(title, enriched_text) and (
            passes_article_filter(title, enriched_text)
            or is_official_ai_law_page(title, enriched_text)
        )
    else:
        relevant = (
            has_explicit_ai_signal(title, summary)
            and passes_article_filter(title, summary)
        )
    if not relevant:
        return None
    article["title"] = title
    article["summary"] = summary
    return attach_analysis(article, title, enriched_text if source_name in OFFICIAL_SOURCE_NAMES else summary)


def load_users() -> dict[str, str]:
    if not USERS_FILE.exists():
        logger.error("Файл users.json не найден. Авторизация невозможна.")
        return {}

    try:
        data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
        users = data.get("users", {})
        if not isinstance(users, dict):
            raise ValueError("Поле users должно быть объектом")
        return {
            str(username).strip(): str(password_hash)
            for username, password_hash in users.items()
            if str(username).strip() and str(password_hash).startswith("pbkdf2_sha256$")
        }
    except Exception as e:
        logger.error(f"Ошибка загрузки пользователей из users.json: {e}")
        return {}


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations, salt, expected = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt.encode("utf-8"),
            int(iterations),
        ).hex()
        return hmac.compare_digest(digest, expected)
    except Exception:
        return False


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if session.get("user") in USERS:
            return view_func(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "authentication_required"}), 401
        return redirect(url_for("login", next=request.path))

    return wrapped


USERS.update(load_users())
DUMMY_PASSWORD_HASH = next(iter(USERS.values()), "pbkdf2_sha256$260000$dummy$"
    "059d8ce6687bd7309efab7af4d7b55557bfc51a7c4cfa9770602cd18f337a04b")


def split_articles_by_mode(articles: list[dict], mode: str) -> list[dict]:
    if mode == "strict":
        return [a for a in articles if a.get("strict_match")]
    if mode == "monitoring":
        return [a for a in articles if not a.get("strict_match")]
    return articles


def article_counts(articles: list[dict]) -> dict:
    strict_count = sum(1 for a in articles if a.get("strict_match"))
    monitoring_count = len(articles) - strict_count
    return {
        "total": len(articles),
        "strict_total": strict_count,
        "monitoring_total": monitoring_count,
    }


def acquire_refresh_lock(owner: str) -> bool:
    with CACHE_LOCK:
        if _cache.get("loading"):
            return False
        if REFRESH_LOCK_FILE.exists():
            try:
                age = time.time() - REFRESH_LOCK_FILE.stat().st_mtime
                if age < REFRESH_LOCK_STALE_SECONDS:
                    return False
                REFRESH_LOCK_FILE.unlink(missing_ok=True)
            except OSError:
                return False
        try:
            descriptor = os.open(
                str(REFRESH_LOCK_FILE),
                os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            )
            with os.fdopen(descriptor, "w", encoding="utf-8") as lock_file:
                lock_file.write(json.dumps({
                    "pid": os.getpid(),
                    "owner": owner,
                    "started_at": datetime.now(timezone.utc).isoformat(),
                }, ensure_ascii=False))
        except FileExistsError:
            return False
        _cache["loading"] = True
        _cache["refresh_owner"] = owner
        return True


def release_refresh_lock() -> None:
    with CACHE_LOCK:
        _cache["loading"] = False
        _cache["refresh_owner"] = None
    try:
        REFRESH_LOCK_FILE.unlink(missing_ok=True)
    except OSError as exc:
        logger.warning(f"Не удалось удалить файл блокировки обновления: {exc}")


def read_history_index() -> list[dict]:
    with HISTORY_LOCK:
        if not HISTORY_INDEX_FILE.exists():
            return []
        try:
            data = json.loads(HISTORY_INDEX_FILE.read_text(encoding="utf-8"))
            return data if isinstance(data, list) else []
        except Exception as e:
            logger.error(f"Ошибка чтения истории обновлений: {e}")
            return []


def write_history_index(history: list[dict]) -> None:
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    temporary = HISTORY_INDEX_FILE.with_suffix(".json.tmp")
    with HISTORY_LOCK:
        temporary.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temporary, HISTORY_INDEX_FILE)


def load_recent_history_articles(limit: int = 5) -> list[dict]:
    """Восстанавливает актуальные материалы из прошлых Excel-отчетов."""
    source_map = {source["name"]: source for source in ALL_SOURCES}
    restored = []
    for record in read_history_index()[:limit]:
        filename = record.get("excel_file", "")
        if not filename:
            continue
        path = HISTORY_DIR / filename
        if not path.exists():
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook["Материалы"]
            headers = {
                cell.value: index
                for index, cell in enumerate(sheet[1])
                if cell.value
            }
            required = {"Дата материала", "Страна/регион", "Категория", "Источник", "Заголовок", "Краткое описание", "URL"}
            if not required.issubset(headers):
                workbook.close()
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = lambda name: row[headers[name]] if headers[name] < len(row) else ""
                title = str(value("Заголовок") or "").strip()
                summary = str(value("Краткое описание") or "").strip()
                link = str(value("URL") or "").strip()
                source_name = str(value("Источник") or "").strip()
                date = parse_date_text(str(value("Дата материала") or ""))
                if not title or not link or not is_recent_date(date) or not is_allowed_source_url(link):
                    continue
                source = source_map.get(source_name, {})
                article = {
                    "id": hashlib.sha256(link.encode("utf-8")).hexdigest()[:24],
                    "title": title,
                    "summary": summary[:500] or "Нет описания",
                    "link": link,
                    "date": date,
                    "collected_at": record.get("created_at"),
                    "source": source_name,
                    "country": str(value("Страна/регион") or source.get("country", "Global")),
                    "flag": source.get("flag", "🌐"),
                    "category": str(value("Категория") or source.get("category", "Правовое регулирование")),
                    "source_description": source.get("description", "Восстановлено из истории обновлений"),
                    "relevance_score": relevance_score(title, summary),
                    "legal_update_type": legal_update_type(title, summary),
                    "legal_priority": legal_priority(title, summary),
                }
                validated = validate_and_analyze_article(article)
                if validated:
                    restored.append(validated)
            workbook.close()
        except Exception as exc:
            logger.warning(f"Не удалось прочитать исторический отчет {path}: {exc}")
    return deduplicate_articles(restored)


def email_status_payload() -> dict:
    config = load_email_config()
    recipients = [r for r in config.get("recipients", []) if r]
    host = config.get("smtp_host") or os.environ.get("AI_MONITOR_SMTP_HOST")
    username = config.get("smtp_username") or os.environ.get("AI_MONITOR_SMTP_USERNAME")
    sender = config.get("from_email") or username
    password = os.environ.get("AI_MONITOR_SMTP_PASSWORD")
    return {
        "enabled": bool(config.get("enabled")),
        "recipients": recipients,
        "recipients_count": len(recipients),
        "smtp_host_set": bool(host),
        "from_email_set": bool(sender),
        "smtp_username_set": bool(username),
        "smtp_password_set": bool(password),
        "ready": bool(config.get("enabled") and recipients and host and sender and username and password),
    }


def config_bool(config: dict, key: str, default: bool = False) -> bool:
    value = config.get(key, default)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def load_email_config() -> dict:
    if not EMAIL_CONFIG_FILE.exists():
        return {"enabled": False, "recipients": []}
    try:
        config = json.loads(EMAIL_CONFIG_FILE.read_text(encoding="utf-8"))
        config.setdefault("enabled", False)
        config.setdefault("recipients", [])
        return config
    except Exception as e:
        logger.error(f"Ошибка загрузки email_config.json: {e}")
        return {"enabled": False, "recipients": []}


def send_email_notification(subject: str, body: str) -> bool:
    config = load_email_config()
    recipients = [r for r in config.get("recipients", []) if r]
    if not config.get("enabled") or not recipients:
        return False

    host = config.get("smtp_host") or os.environ.get("AI_MONITOR_SMTP_HOST")
    port = int(config.get("smtp_port") or os.environ.get("AI_MONITOR_SMTP_PORT", "587"))
    username = config.get("smtp_username") or os.environ.get("AI_MONITOR_SMTP_USERNAME")
    password = os.environ.get("AI_MONITOR_SMTP_PASSWORD")
    sender = config.get("from_email") or username
    use_tls = config_bool(config, "use_tls", True)
    use_ssl = config_bool(config, "use_ssl", False) or config_bool(config, "smtp_ssl", False)
    timeout = int(config.get("timeout") or os.environ.get("AI_MONITOR_SMTP_TIMEOUT", "45"))
    if not host or not sender:
        logger.warning("Email не отправлен: не указан SMTP host/from_email")
        return False

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)

    try:
        context = ssl.create_default_context()
        if use_ssl:
            smtp_server = smtplib.SMTP_SSL(host, port, timeout=timeout, context=context)
        else:
            smtp_server = smtplib.SMTP(host, port, timeout=timeout)
        with smtp_server as smtp:
            smtp.ehlo()
            if use_tls and not use_ssl:
                smtp.starttls(context=context)
                smtp.ehlo()
            if username and password:
                smtp.login(username, password)
            smtp.send_message(msg)
        return True
    except Exception as e:
        logger.error(f"Ошибка отправки email: {e}")
        return False


def normalize_article_link(link: str) -> str:
    if not link:
        return ""
    try:
        parsed = urllib.parse.urlparse(link.strip())
        query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
        ignored = {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "fbclid", "gclid"}
        if parsed.netloc.lower() == "eur-lex.europa.eu":
            ignored.update({"qid", "rid"})
        query = [(k, v) for k, v in query if k.lower() not in ignored]
        return urllib.parse.urlunparse((
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path.rstrip("/"),
            "",
            urllib.parse.urlencode(query),
            "",
        ))
    except Exception:
        return link.strip()


def diagnostic_brief(diagnostics: list[dict]) -> list[str]:
    if not diagnostics:
        return ["Диагностика источников: нет данных."]
    failed = [d for d in diagnostics if d.get("status") in {"failed", "timeout", "error"}]
    unavailable = [d for d in diagnostics if d.get("status") == "unavailable"]
    accepted = [d for d in diagnostics if int(d.get("accepted_count") or 0) > 0]
    lines = [
        (
            f"Источники: проверено {len(diagnostics)}, дали материалы {len(accepted)}, "
            f"с ошибками/таймаутами {len(failed)}, недоступных/заблокированных {len(unavailable)}."
        ),
    ]
    for item in failed[:8]:
        lines.append(f"- {item.get('name', 'source')}: {item.get('status')} — {item.get('message', '')}")
    if unavailable:
        lines.append("Недоступные источники обычно означают 403/404, отключенный RSS или блокировку серверных запросов.")
    return lines


def article_email_lines(articles: list[dict]) -> list[str]:
    lines: list[str] = []
    for article in articles[:25]:
        mode = "строгий режим" if article.get("strict_match") else "расширенный режим"
        lines.extend([
            f"[{mode}] {article.get('title', 'Без заголовка')}",
            f"{article.get('summary', '')[:260]}",
            f"URL: {article.get('link', '')}",
            "",
        ])
    return lines


def notify_about_refresh(previous_links: set[str], articles: list[dict], diagnostics: list[dict] | None = None) -> None:
    diagnostics = diagnostics or []
    new_articles = [
        a for a in articles
        if a.get("link") and normalize_article_link(a.get("link", "")) not in previous_links
    ]

    if new_articles:
        lines = [
            "Еженедельная проверка выполнена.",
            f"Новых материалов: {len(new_articles)}.",
            "",
            "Новые материалы по ИИ и правовому регулированию:",
            "",
        ]
        lines.extend(article_email_lines(new_articles))
        lines.extend(["", *diagnostic_brief(diagnostics)])
        send_email_notification("AI Regulation Monitor: новые материалы", "\n".join(lines))
        return

    lines = [
        "Еженедельная проверка выполнена.",
        "Новых материалов по ИИ и правовому регулированию не найдено.",
        f"Всего актуальных материалов в базе: {len(articles)}.",
        "",
        *diagnostic_brief(diagnostics),
    ]
    send_email_notification("AI Regulation Monitor: новых материалов нет", "\n".join(lines))


def notify_about_new_articles(previous_links: set[str], articles: list[dict]) -> None:
    notify_about_refresh(previous_links, articles, [])


def human_status(status: str) -> str:
    labels = {
        "ok": "Успешно",
        "no-records": "Нет релевантных записей",
        "unavailable": "Недоступен",
        "pending": "Ожидает",
        "running": "В работе",
        "timeout": "Таймаут",
        "error": "Ошибка",
        "failed": "Ошибка",
    }
    return labels.get(status or "", status or "Неизвестно")


def auto_fit_sheet(sheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        sheet.column_dimensions[chr(64 + col_idx)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="8B1A1A")
    border = Border(bottom=Side(style="thin", color="D9D4CC"))
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_excel_report(articles: list[dict], diagnostics: list[dict], updated_at: str, file_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Материалы"
    headers = [
        "№",
        "Режим",
        "Дата материала",
        "Страна/регион",
        "Категория",
        "Тип правового события",
        "Источник",
        "Заголовок",
        "Краткое описание",
        "URL",
        "NLP-метка",
        "NLP уверенность",
        "Причина отбора",
        "Также найдено в источниках",
    ]
    ws.append(headers)
    for index, article in enumerate(articles, start=1):
        title = article.get("title", "")
        summary = article.get("summary", "")
        url = article.get("link", "")
        row = [
            index,
            "Строгий режим" if article.get("strict_match") else "Расширенный режим",
            article.get("date", ""),
            article.get("country", ""),
            article.get("category", ""),
            legal_update_type(title, summary),
            article.get("source", ""),
            title,
            summary,
            url,
            article.get("nlp_label") or article.get("filter_mode", ""),
            article.get("nlp_confidence") or article.get("analysis_confidence", ""),
            article.get("nlp_reason") or article.get("analysis_reason", ""),
            ", ".join(article.get("also_reported_by", [])),
        ]
        ws.append([neutralize_spreadsheet_value(value) for value in row])
        if url:
            url_cell = ws.cell(row=index + 1, column=10)
            url_cell.hyperlink = url
            url_cell.style = "Hyperlink"
    style_header(ws)
    auto_fit_sheet(ws, {
        1: 6, 2: 18, 3: 18, 4: 16, 5: 24, 6: 26, 7: 32,
        8: 52, 9: 72, 10: 64, 11: 16, 12: 16, 13: 72, 14: 48,
    })

    ds = wb.create_sheet("Диагностика")
    ds.append([
        "Источник",
        "Страна",
        "Статус",
        "Найдено",
        "Принято",
        "Строгих",
        "Расширенных",
        "Отклонено",
        "Сообщение",
        "URL",
    ])
    for item in diagnostics:
        status = item.get("status", "")
        ds.append([
            item.get("name", ""),
            item.get("country", ""),
            human_status(status),
            item.get("found_count", 0),
            item.get("accepted_count", 0),
            item.get("strict_count", 0),
            item.get("monitoring_count", 0),
            item.get("rejected_count", 0),
            item.get("message", ""),
            item.get("url", ""),
        ])
        url = item.get("url", "")
        if url:
            url_cell = ds.cell(row=ds.max_row, column=10)
            url_cell.hyperlink = url
            url_cell.style = "Hyperlink"
    style_header(ds)
    auto_fit_sheet(ds, {1: 36, 2: 16, 3: 20, 4: 12, 5: 12, 6: 12, 7: 16, 8: 12, 9: 72, 10: 64})

    ss = wb.create_sheet("Сводка")
    counts = article_counts(articles)
    stats = get_stats(articles)
    ok_sources = sum(1 for d in diagnostics if d.get("status") == "ok")
    problem_sources = len(diagnostics) - ok_sources
    summary_rows = [
        ["Показатель", "Значение", "Комментарий"],
        ["Время обновления", updated_at, "Фиксируется при завершении сбора данных"],
        ["Всего материалов", counts["total"], "Строгий + расширенный режим"],
        ["Строгий режим", counts["strict_total"], "Подтвержденные правовые акты, проекты, изменения, вступление в силу"],
        ["Расширенный режим", counts["monitoring_total"], "Правовая повестка, планы, обсуждения и новости о регулировании ИИ"],
        ["Источников в конфигурации", len(ALL_SOURCES), "Официальные и профильные источники"],
        ["Источников без ошибок", ok_sources, ""],
        ["Источников с предупреждениями/ошибками", problem_sources, ""],
    ]
    for row in summary_rows:
        ss.append(row)
    ss.append([])
    ss.append(["Страна/регион", "Материалов", ""])
    for country, value in stats.get("countries", {}).items():
        ss.append([country, value, ""])
    ss.append([])
    ss.append(["Категория", "Материалов", ""])
    for category, value in stats.get("categories", {}).items():
        ss.append([category, value, ""])
    style_header(ss)
    for row_idx in (10 + len(stats.get("countries", {})) + 1,):
        for cell in ss[row_idx]:
            cell.font = Font(bold=True, color="8B1A1A")
    auto_fit_sheet(ss, {1: 32, 2: 24, 3: 76})

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)


def save_update_history(articles: list[dict], diagnostics: list[dict], auto_refresh: bool = False) -> dict | None:
    try:
        HISTORY_DIR.mkdir(parents=True, exist_ok=True)
        created_at = datetime.now(timezone.utc).isoformat()
        record_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
        filename = f"update_{record_id}.xlsx"
        file_path = HISTORY_DIR / filename
        build_excel_report(articles, diagnostics, created_at, file_path)
        counts = article_counts(articles)
        ok_sources = sum(1 for d in diagnostics if d.get("status") == "ok")
        record = {
            "id": record_id,
            "created_at": created_at,
            "total": counts["total"],
            "strict_total": counts["strict_total"],
            "monitoring_total": counts["monitoring_total"],
            "sources_count": len(ALL_SOURCES),
            "diagnostics_total": len(diagnostics),
            "ok_sources": ok_sources,
            "problem_sources": max(0, len(diagnostics) - ok_sources),
            "excel_file": filename,
            "auto_refresh": auto_refresh,
        }
        with HISTORY_LOCK:
            history = [item for item in read_history_index() if item.get("id") != record_id]
            history.insert(0, record)
            write_history_index(history[:100])
        logger.info(f"Отчёт обновления сохранён: {file_path}")
        return record
    except Exception as e:
        logger.error(f"Ошибка сохранения истории обновлений: {e}")
        return None


def save_cache(articles: list, diagnostics: list | None = None, last_auto_refresh: str | None = None) -> None:
    global _cache_file_mtime
    payload = {
        "cache_version": CACHE_VERSION,
        "articles": articles,
        "stats": get_stats(articles),
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "sources_count": len(ALL_SOURCES),
        "diagnostics": diagnostics or [],
        "last_auto_refresh": last_auto_refresh or _cache.get("last_auto_refresh"),
    }
    with CACHE_LOCK:
        _cache.update(payload)
    try:
        CACHE_TEMP_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(CACHE_TEMP_FILE, CACHE_FILE)
        _cache_file_mtime = CACHE_FILE.stat().st_mtime
        logger.info("Кэш сохранён на диск")
    except Exception as e:
        logger.error(f"Ошибка сохранения кэша: {e}")


def load_cache_from_disk() -> bool:
    global _cache_file_mtime
    if not CACHE_FILE.exists():
        return False
    try:
        data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        if data.get("cache_version") != CACHE_VERSION or data.get("sources_count") != len(ALL_SOURCES):
            logger.info("Кэш устарел после изменения источников. Требуется ручное обновление данных.")
            return False
        articles = []
        for article in data.get("articles", []):
            article = validate_and_analyze_article(article)
            if not article:
                continue
            if "прямой поиск" in article.get("source", "") and not article.get("collected_at"):
                article["collected_at"] = article.get("date")
                article["date"] = parse_date_text(article.get("title", ""), article.get("summary", ""))
            if not is_recent_date(article.get("date", "")):
                continue
            article.setdefault(
                "legal_update_type",
                legal_update_type(article.get("title", ""), article.get("summary", "")),
            )
            article.setdefault(
                "legal_priority",
                legal_priority(article.get("title", ""), article.get("summary", "")),
            )
            articles.append(article)
        articles.sort(
            key=lambda a: (
                to_datetime(a.get("date", "")),
                int(a.get("legal_priority", 1)),
                int(a.get("relevance_score", 0)),
            ),
            reverse=True,
        )
        data["articles"] = articles
        data["stats"] = get_stats(articles)
        data["sources_count"] = len(ALL_SOURCES)
        data.setdefault("diagnostics", [])
        with CACHE_LOCK:
            _cache.update(data)
            _cache["loading"] = False
            _cache["refresh_owner"] = None
        _cache_file_mtime = CACHE_FILE.stat().st_mtime
        logger.info(f"Кэш загружен с диска: {len(data.get('articles', []))} статей")
        return True
    except Exception as e:
        logger.error(f"Ошибка загрузки кэша: {e}")
        return False


def sync_cache_from_disk_if_changed() -> None:
    if not CACHE_FILE.exists():
        return
    try:
        disk_mtime = CACHE_FILE.stat().st_mtime
        with CACHE_LOCK:
            loading = bool(_cache.get("loading"))
        if not loading and disk_mtime > _cache_file_mtime:
            load_cache_from_disk()
    except OSError:
        return



def refresh_articles(
    send_notifications: bool = False,
    auto_refresh: bool = False,
    owner: str = "system",
    lock_acquired: bool = False,
) -> None:
    if not lock_acquired and not acquire_refresh_lock(owner):
        logger.info("Обновление не запущено: другой сбор уже выполняется")
        return
    logger.info("▶ Начало обновления данных...")
    with CACHE_LOCK:
        _cache["refresh_status"] = "running"
        _cache["refresh_error"] = None
        _cache["refresh_finished_at"] = None
        previous_articles = list(_cache.get("articles", []))
        previous_links = {
            normalize_article_link(a.get("link", ""))
            for a in previous_articles
            if a.get("link")
        }
    try:
        result = fetch_all_articles(include_diagnostics=True)
        fresh_articles = result.get("articles", []) if isinstance(result, dict) else result
        diagnostics = result.get("diagnostics", []) if isinstance(result, dict) else []
        retained_articles = [
            validated
            for article in previous_articles
            if is_recent_date(article.get("date", ""))
            for validated in [validate_and_analyze_article(article)]
            if validated
        ]
        history_articles = load_recent_history_articles()
        articles = deduplicate_articles(
            list(fresh_articles) + retained_articles + history_articles
        )
        articles.sort(
            key=lambda article: (
                to_datetime(article.get("date", "")),
                int(article.get("legal_priority", 1)),
                int(article.get("relevance_score", 0)),
            ),
            reverse=True,
        )
        logger.info(
            "Скользящий архив: новых/повторно найденных %s, из кэша %s, из истории %s, итог %s",
            len(fresh_articles),
            len(retained_articles),
            len(history_articles),
            len(articles),
        )
        refresh_time = datetime.now(timezone.utc).isoformat() if auto_refresh else _cache.get("last_auto_refresh")
        save_cache(articles, diagnostics, last_auto_refresh=refresh_time)
        save_update_history(articles, diagnostics, auto_refresh=auto_refresh)
        if auto_refresh:
            with CACHE_LOCK:
                _cache["last_auto_refresh"] = refresh_time
        if send_notifications:
            notify_about_refresh(previous_links, articles, diagnostics)
        with CACHE_LOCK:
            _cache["refresh_status"] = "success"
            _cache["refresh_error"] = None
            _cache["refresh_finished_at"] = datetime.now(timezone.utc).isoformat()
        logger.info(f"✅ Обновление завершено: {len(articles)} статей")
    except Exception as e:
        with CACHE_LOCK:
            _cache["refresh_status"] = "failed"
            _cache["refresh_error"] = str(e)[:300]
            _cache["refresh_finished_at"] = datetime.now(timezone.utc).isoformat()
        logger.exception("Ошибка при обновлении")
    finally:
        release_refresh_lock()


def background_refresh(
    send_notifications: bool = False,
    auto_refresh: bool = False,
    owner: str = "system",
    lock_acquired: bool = False,
):
    t = threading.Thread(
        target=refresh_articles,
        kwargs={
            "send_notifications": send_notifications,
            "auto_refresh": auto_refresh,
            "owner": owner,
            "lock_acquired": lock_acquired,
        },
        daemon=True,
    )
    t.start()


def weekly_refresh_loop() -> None:
    while True:
        time.sleep(WEEKLY_REFRESH_INTERVAL_SECONDS)
        if not acquire_refresh_lock("weekly"):
            logger.info("Еженедельное обновление пропущено: уже идет сбор")
            continue
        refresh_articles(
            send_notifications=True,
            auto_refresh=True,
            owner="weekly",
            lock_acquired=True,
        )


def start_weekly_refresh_thread() -> None:
    thread = threading.Thread(target=weekly_refresh_loop, daemon=True)
    thread.start()



@app.route("/robots.txt")
def robots_txt():
    return "User-agent: *\nDisallow: /\n", 200, {"Content-Type": "text/plain; charset=utf-8"}


@app.route("/")
@login_required
def index():
    return render_template("index.html", username=session.get("user"), csrf_token=csrf_token())


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user") in USERS:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        if not csrf_is_valid():
            return render_template("login.html", error="Сессия формы устарела. Обновите страницу.", csrf_token=csrf_token()), 400
        if login_rate_limited():
            return render_template("login.html", error="Слишком много попыток. Повторите позже.", csrf_token=csrf_token()), 429
        username = request.form.get("username", "").strip()[:128]
        password = request.form.get("password", "")[:1024]
        stored_hash = USERS.get(username, DUMMY_PASSWORD_HASH)
        password_ok = verify_password(password, stored_hash)
        if username in USERS and password_ok:
            clear_login_failures()
            next_url = safe_next_url(request.args.get("next"))
            session.clear()
            session["user"] = username
            csrf_token()
            return redirect(next_url)
        record_login_failure()
        error = "Неверный логин или пароль"

    return render_template("login.html", error=error, csrf_token=csrf_token())


@app.route("/logout", methods=["POST"])
def logout():
    if not csrf_is_valid():
        return jsonify({"error": "invalid_csrf_token"}), 403
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/status")
@login_required
def api_status():
    sync_cache_from_disk_if_changed()
    with CACHE_LOCK:
        articles = list(_cache.get("articles", []))
        counts = article_counts(articles)
        return jsonify({
            "loading": _cache.get("loading", False),
            "last_updated": _cache.get("last_updated"),
            "total": counts["total"],
            "strict_total": counts["strict_total"],
            "monitoring_total": counts["monitoring_total"],
            "sources_count": _cache.get("sources_count", 0),
            "last_auto_refresh": _cache.get("last_auto_refresh"),
            "refresh_owner": _cache.get("refresh_owner"),
            "refresh_status": _cache.get("refresh_status", "idle"),
            "refresh_error": _cache.get("refresh_error"),
            "refresh_finished_at": _cache.get("refresh_finished_at"),
        })


@app.route("/api/articles")
@login_required
def api_articles():
    sync_cache_from_disk_if_changed()
    country = request.args.get("country", "")
    category = request.args.get("category", "")
    query = request.args.get("q", "").lower().strip()
    mode = request.args.get("mode", "strict")
    if mode not in {"strict", "monitoring"}:
        return jsonify({"error": "invalid_mode"}), 400
    try:
        page = max(1, int(request.args.get("page", 1)))
        per_page = int(request.args.get("per_page", 20))
    except (TypeError, ValueError):
        return jsonify({"error": "invalid_pagination"}), 400
    if per_page < 1 or per_page > 50:
        return jsonify({"error": "invalid_per_page", "min": 1, "max": 50}), 400

    with CACHE_LOCK:
        articles = split_articles_by_mode(list(_cache.get("articles", [])), mode)

    # Фильтрация
    if country:
        articles = [a for a in articles if a.get("country") == country]
    if category:
        articles = [a for a in articles if a.get("category") == category]
    if query:
        articles = [
            a for a in articles
            if query in a.get("title", "").lower()
            or query in a.get("summary", "").lower()
        ]

    total = len(articles)
    start = (page - 1) * per_page
    end = start + per_page
    page_articles = articles[start:end]

    return jsonify({
        "articles": page_articles,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, -(-total // per_page)),  # ceil division
    })


@app.route("/api/stats")
@login_required
def api_stats():
    sync_cache_from_disk_if_changed()
    mode = request.args.get("mode", "strict")
    if mode not in {"strict", "monitoring"}:
        return jsonify({"error": "invalid_mode"}), 400
    with CACHE_LOCK:
        all_articles = list(_cache.get("articles", []))
        counts = article_counts(all_articles)
        articles = split_articles_by_mode(all_articles, mode)
        stats = get_stats(articles)
        all_stats = get_stats(all_articles)
        countries = dict(stats.get("countries", {}))
        categories = dict(stats.get("categories", {}))
        all_countries = dict(all_stats.get("countries", {}))
        all_categories = dict(all_stats.get("categories", {}))
        for source in ALL_SOURCES:
            countries.setdefault(source["country"], 0)
            categories.setdefault(source["category"], 0)
            all_countries.setdefault(source["country"], 0)
            all_categories.setdefault(source["category"], 0)
        stats["countries"] = dict(sorted(countries.items(), key=lambda x: (-x[1], x[0])))
        stats["categories"] = dict(sorted(categories.items(), key=lambda x: (-x[1], x[0])))
        all_stats["countries"] = dict(sorted(all_countries.items(), key=lambda x: (-x[1], x[0])))
        all_stats["categories"] = dict(sorted(all_categories.items(), key=lambda x: (-x[1], x[0])))
        return jsonify({
            "stats": stats,
            "all_stats": all_stats,
            "total_all": counts["total"],
            "strict_total": counts["strict_total"],
            "monitoring_total": counts["monitoring_total"],
            "last_updated": _cache.get("last_updated"),
            "loading": _cache.get("loading", False),
        })


def sanitize_diagnostic_message(message: str) -> str:
    text = clean_html(str(message or "")).strip()
    lowered = text.lower()
    if not text:
        return "Источник проверен; дополнительных сведений нет."
    if "403" in lowered or "forbidden" in lowered:
        return "Источник отклонил автоматический запрос (HTTP 403). Использован резервный способ, если он настроен."
    if "404" in lowered or "not found" in lowered:
        return "Страница источника не найдена (HTTP 404); адрес требует проверки."
    if "410" in lowered:
        return "Страница источника удалена (HTTP 410); адрес требует замены."
    if "timeout" in lowered or "timed out" in lowered:
        return "Источник не ответил за отведенное время. Проверка продолжилась без остановки всего обновления."
    if "connection" in lowered or "max retries" in lowered or "ssl" in lowered:
        return "Не удалось установить защищенное соединение с источником."
    if "empty or broken feed" in lowered:
        return "Лента источника пуста или имеет неподдерживаемый формат."
    return text[:300]


@app.route("/api/diagnostics")
@login_required
def api_diagnostics():
    sync_cache_from_disk_if_changed()
    source_names = {source["name"] for source in ALL_SOURCES}
    with CACHE_LOCK:
        diagnostics = [
            dict(item)
            for item in _cache.get("diagnostics", [])
            if item.get("name") in source_names
        ]
    if not diagnostics:
        diagnostics = [
            {
                "name": s["name"],
                "status": "pending",
                "raw_count": 0,
                "candidate_count": 0,
                "accepted_count": 0,
                "strict_count": 0,
                "monitoring_count": 0,
                "message": "Диагностика появится после обновления данных",
                "url": s.get("url", ""),
            }
            for s in ALL_SOURCES
        ]
    for item in diagnostics:
        message = item.get("message", "")
        item["message"] = sanitize_diagnostic_message(message)
        if item.get("status") == "failed" and any(token in message for token in ("HTTP 403", "HTTP 404", "HTTP 410", "Forbidden", "Not Found", "empty or broken feed")):
            item["status"] = "unavailable"
        elif item.get("status") == "ok" and int(item.get("accepted_count", 0)) == 0:
            item["status"] = "no-records"
    status_rank = {"ok": 0, "no-records": 1, "unavailable": 2, "pending": 3, "running": 4, "timeout": 5, "error": 6, "failed": 7}
    diagnostics.sort(key=lambda item: (
        status_rank.get(item.get("status", "ok"), 8),
        -int(item.get("accepted_count", 0)),
        item.get("name", ""),
    ))
    return jsonify({"diagnostics": diagnostics})


@app.route("/api/sources")
@login_required
def api_sources():
    return jsonify({
        "sources": [
            {
                "name": s["name"],
                "country": s["country"],
                "flag": s["flag"],
                "category": s["category"],
                "description": s["description"],
            }
            for s in ALL_SOURCES
        ]
    })


@app.route("/api/refresh", methods=["POST"])
@login_required
def api_refresh():
    if not csrf_is_valid():
        return jsonify({"error": "invalid_csrf_token"}), 403
    owner = session.get("user", "manual")
    if not acquire_refresh_lock(owner):
        return jsonify({"status": "already_loading"}), 200
    background_refresh(
        send_notifications=True,
        owner=owner,
        lock_acquired=True,
    )
    return jsonify({"status": "started"})


@app.route("/api/history")
@login_required
def api_history():
    history = read_history_index()
    return jsonify({"history": history})


@app.route("/api/history/<record_id>/download")
@login_required
def api_history_download(record_id: str):
    history = read_history_index()
    record = next((item for item in history if item.get("id") == record_id), None)
    if not record:
        return jsonify({"error": "history_record_not_found"}), 404
    filename = record.get("excel_file", "")
    if not filename or "/" in filename or "\\" in filename:
        return jsonify({"error": "invalid_report_file"}), 400
    file_path = HISTORY_DIR / filename
    if not file_path.exists():
        return jsonify({"error": "report_file_not_found"}), 404
    return send_from_directory(str(HISTORY_DIR.resolve()), filename, as_attachment=True)


@app.route("/api/email/status")
@login_required
def api_email_status():
    return jsonify(email_status_payload())


@app.route("/api/filters")
@login_required
def api_filters():
    sync_cache_from_disk_if_changed()
    with CACHE_LOCK:
        articles = _cache.get("articles", [])
    countries = sorted({s["country"] for s in ALL_SOURCES} | {a["country"] for a in articles})
    categories = sorted({s["category"] for s in ALL_SOURCES} | {a["category"] for a in articles})
    return jsonify({"countries": countries, "categories": categories})



if __name__ == "__main__":
    if not USERS:
        logger.error("Нет доступных учетных записей. Заполните users.json перед запуском.")

    has_cache = load_cache_from_disk()

    if not has_cache:
        logger.info("Кэш не найден. Автоматический сбор отключен — нажмите «Обновить данные» в интерфейсе.")

    if os.environ.get("AI_MONITOR_ENABLE_SCHEDULER", "1").lower() in {"1", "true", "yes"}:
        start_weekly_refresh_thread()
    threading.Timer(1.5, lambda: webbrowser.open_new_tab("http://127.0.0.1:5000")).start()    
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False, threaded=True)
